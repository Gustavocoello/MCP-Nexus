# src/services/llm/providers/utils.py
import os
import time
import magic
import pypdf
import docx2txt
from re import A 
from flask import g
from io import BytesIO
from openai import OpenAI
from datetime import datetime, timezone
from dotenv import load_dotenv
from openpyxl import load_workbook
from src.config.time_helper import get_now
from src.database.models.models import Document
from src.config.logging_config import get_logger 
from src.database.config.connection import SessionLocal 
from azure.cognitiveservices.vision.computervision import ComputerVisionClient
from azure.cognitiveservices.vision.computervision.models import VisualFeatureTypes
from msrest.authentication import CognitiveServicesCredentials

# Configuración de logging
logger = get_logger(__name__)

# Cargar las variables de entorno desde el archivo .env
load_dotenv()

OPENROUTER0 = os.getenv("OPEN_ROUTER_0")
OPENROUTER1 = os.getenv("OPEN_ROUTER_1")
OPENROUTER2 = os.getenv("OPEN_ROUTER_2")


AZURE_ENDPOINT = os.getenv("AZURE_VISION_ENDPOINT")
AZURE_KEY = os.getenv("AZURE_KEY1")


API_PROVIDERS = [
    {
        'name'     : 'Kwaipilot',
        'client'   : OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key= OPENROUTER0),
        'handler'  : lambda client, prompt: client.chat.completions.create(
            model = 'kwaipilot/kat-coder-pro:free',
            messages = [{'role' : 'user', 'content' : prompt}]
        )
    },
    {
        'name'     : 'Nemotron Nano 12B',
        'client'   : OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key= OPENROUTER1),
        'handler'  : lambda client, prompt: client.chat.completions.create(
            model = 'nvidia/nemotron-nano-12b-v2-vl:free',
            messages = [{'role' : 'user', 'content' : prompt}]
        )
    },
    {
        'name'     : 'TNG: R1T Chimera',
        'client'   : OpenAI(
            base_url="https://openrouter.ai/api/v1",
            api_key= OPENROUTER1),
        'handler'  : lambda client, prompt: client.chat.completions.create(
            model = 'tngtech/tng-r1t-chimera:free',
            messages = [{'role' : 'user', 'content' : prompt}]
        )
    }    
    # Se puede añadir más apis aqui para abajo  ->
]

# DeepSeek índice = 0
# Qwen3 índice = 1
# OpenAI índice = 2
# Gemini índice = 3

current_provider_index = 0
MAX_RETRIES = 2 

def generate_prompt(prompt_data, retry_count=0):
    
    logger.debug(f"Modo prueba activado: {os.getenv('TEST_MODE')}")
    
    global current_provider_index
    
    if current_provider_index >= len(API_PROVIDERS):
        return "Todos los servicios de IA no estan disponibles actualmente."
    
    provider = API_PROVIDERS[current_provider_index]
    logger.info(f"usando el {provider['name']} proveedor (Intento {retry_count + 1} /{MAX_RETRIES})")
    
    try:
        # ======================================
        # SIMULADOR DE ERRORES SOLO PARA PRUEBAS
        # ======================================
        if provider["name"] == 'Kwaipilot' and os.getenv('TEST_MODE') == 'True':
            if retry_count == 0:
                raise ConnectionError("Error de conexion simulado con Kwaipilot")
            elif retry_count == 1:
                raise Exception("Error de timeout simulado con Kwaipilot")
            
        user_prompt = prompt_data.get("description", "") # actualizado
        prompt = f"{user_prompt}"
        
        response = provider['handler'](provider['client'], prompt)
        return response.choices[0].message.content 
       
            
    except Exception as e:
        logger.error(f"Error with {provider['name']}: {str(e)}")
        
        """
        Casos que activarán el cambio de proveedor:
            - Errores de conexión (timeout, red caída)
            - Límites de tasa excedidos (rate limits)
            - Errores de autenticación (API key inválida)
            - Errores internos del servidor (5xx)
            - Cualquier excepción no manejada
        """
        
        if retry_count < MAX_RETRIES -1:
            return generate_prompt(prompt_data, retry_count + 1)
        else:
            # Cambiar al siguiente proveedor
            current_provider_index += 1
            return generate_prompt(prompt_data, 0)
        


# =============== Función para extraer texto de archivos ==================
def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def extract_text_from_file(file_stream, filename):
    file_stream.seek(0)
    file_type = magic.from_buffer(file_stream.read(1024), mime=True)
    file_stream.seek(0)

    logger.info(f"Procesando archivo: {filename} con MIME: {file_type}")

    # PDF
    if 'pdf' in file_type or filename.endswith('.pdf'):
        reader = pypdf.PdfReader(file_stream)
        text = "\n".join([page.extract_text() or '' for page in reader.pages])
    
    # Word moderno
    elif 'word' in file_type or filename.endswith('.docx'):
        text = docx2txt.process(file_stream)

    # Excel
    elif 'excel' in file_type or filename.endswith(('.xlsx', '.xls')):
        file_stream.seek(0)

        # openpyxl solo soporta .xlsx.
        if filename.endswith('.xlsx'):
            workbook = load_workbook(filename=BytesIO(file_stream.read()))
            text = ""
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    text += "\t".join(str(cell.value) for cell in row if cell.value is not None) + "\n"
        else:
            # Mantener xlrd para .xls antiguos, o usar una librería diferente (ej. pyxlsb)
            # Para este ejemplo, si solo quieres .xlsx, podrías eliminar el soporte .xls
            # o dejar la lógica actual de xlrd para .xls.
            raise ValueError("Soporte para .xls requiere xlrd; se usa openpyxl para .xlsx")

    # Texto plano
    elif file_type.startswith('text') or filename.endswith(('.txt', '.csv')):
        content = file_stream.read().decode('utf-8', errors='ignore')
        text = content

    # Archivos no soportados
    else:
        logger.warning(f"Tipo de archivo no soportado: {filename} ({file_type})")
        raise ValueError("Formato de archivo no soportado")

    logger.info(f"Texto extraído de {filename} (longitud: {len(text)} caracteres)")
    return text



# ====================== Funcion para extraer texto de imagenes Azure AI Vision========================
OCR_LIMIT = 5000

def analyze_image_with_azure(image_bytes):
    endpoint = AZURE_ENDPOINT
    key = AZURE_KEY

    if not endpoint or not key:
        raise ValueError("Azure Vision endpoint o key no definidos.")

    client = ComputerVisionClient(endpoint, CognitiveServicesCredentials(key))
    stream = BytesIO(image_bytes)

    # --- Control de límite mensual OCR ---
    now = get_now()
    first_day_of_month = datetime(now.year, now.month, 1)
    
    # Contar cuántas imágenes ha subido el usuario este mes
    session = SessionLocal()
    ocr_count = session.query(Document).filter(
        Document.user_id == g.user_id,
        Document.created_at >= first_day_of_month,
        Document.source == "onedrive"  # solo imágenes subidas a OneDrive
    ).count()
    session.close()

    if ocr_count >= OCR_LIMIT:
        raise ValueError("Límite mensual de OCR alcanzado (5000 imágenes).")

    # Asignamos el número OCR incremental
    ocr_number = ocr_count + 1

    # --- Parte 1: Análisis Visual ---
    analysis = client.analyze_image_in_stream(
        stream,
        visual_features=[VisualFeatureTypes.description,
                         VisualFeatureTypes.color,
                         VisualFeatureTypes.tags]
    )

    description = analysis.description.captions[0].text if analysis.description.captions else "Sin descripción."
    colors = {
        "fondo": analysis.color.dominant_color_background,
        "frente": analysis.color.dominant_color_foreground,
        "acentos": analysis.color.accent_color
    }
    tags = [tag.name for tag in analysis.tags][:5]

    visual_summary = f""" **Imagen #{ocr_number}** (proporcionada por usuario)
        **Descripción**: {description}
        **Colores**: Fondo: {colors['fondo']}, Frente: {colors['frente']}, Acento: #{colors['acentos']}
        **Etiquetas**: {', '.join(tags)}
        """

    # --- Parte 2: OCR ---
    ocr_stream = BytesIO(image_bytes)
    result = client.read_in_stream(ocr_stream, raw=True)
    operation_location = result.headers["Operation-Location"]
    operation_id = operation_location.split("/")[-1]

    while True:
        read_result = client.get_read_result(operation_id)
        if read_result.status.lower() in ["succeeded", "failed"]:
            break
        time.sleep(0.5)

    if read_result.status.lower() != "succeeded":
        raise ValueError("Error procesando imagen con Azure Vision OCR")

    ocr_lines = []
    for page in read_result.analyze_result.read_results:
        for line in page.lines:
            ocr_lines.append(line.text)

    ocr_text = "\n".join(ocr_lines) if ocr_lines else "Sin texto detectado."
    
    # --- Unión final ---
    full_context = f"{visual_summary}\n**Texto OCR:**\n{ocr_text}"

    return full_context


def can_upload_image(username: str) -> bool:
    """
    Retorna True si el usuario puede subir imágenes.
    Solo permite usuarios cuyo nombre empieza con 'G' o 'U'.
    """
    if not username:
        return False
    return username.upper().startswith(("G", "U"))
